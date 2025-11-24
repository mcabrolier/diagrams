import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
OUTPUT_FILE = BASE_DIR / "Portfolio_Convergence.xlsx"

def read_csv(name: str) -> pd.DataFrame:
    return pd.read_csv(BASE_DIR / name)

def compute_kpi(products: pd.DataFrame, scores: pd.DataFrame) -> pd.DataFrame:
    # TCI par produit / snapshot
    grouped = scores.groupby(["snapshot_date", "product_id"]).agg(
        total_weighted=("score_weighted", "sum"),
        total_criteria_weight=("criterion_weight", "sum"),
        raw_scores=("score", list)
    ).reset_index()
    grouped["tci"] = grouped.apply(
        lambda r: 0 if r.total_criteria_weight == 0 else r.total_weighted / r.total_criteria_weight,
        axis=1
    )
    grouped["tci_pct"] = grouped["tci"] / 5 * 100

    # SAI: % critères >=4
    def sai(sub):
        return (sub["score"] >= 4).sum() / len(sub) * 100 if len(sub) else 0
    sai_df = scores.groupby(["snapshot_date", "product_id"]).apply(sai).reset_index(name="sai_pct")

    # Merge
    kpi = grouped.merge(sai_df, on=["snapshot_date", "product_id"])

    # REI (risque simplifié) – sera calculé plus tard si besoin, placeholder
    kpi["rei_open_high"] = 0  # remplacé après jointure dette

    return kpi

def integrate_debt(kpi: pd.DataFrame, debt: pd.DataFrame) -> pd.DataFrame:
    # Compte dettes high non done par produit/snapshot (approximation : snapshot_date max affecte)
    high_open = debt[debt["risk_level"] == "high"].query("status != 'done'")
    counts = high_open.groupby("product_id").size().reset_index(name="open_high_debts")
    return kpi.merge(counts, on="product_id", how="left").fillna({"open_high_debts": 0})

def format_sheet(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def main():
    products = read_csv("Products.csv")
    criteria = read_csv("Criteria.csv")
    scores = read_csv("Scores.csv")
    debt = read_csv("DebtRegister.csv")

    kpi = compute_kpi(products, scores)
    kpi = integrate_debt(kpi, debt)

    # Ajout criticity_weight
    kpi = kpi.merge(products[["product_id", "criticality_weight"]], on="product_id", how="left")
    kpi["weighted_tci_portfolio_component"] = kpi["tci"] * kpi["criticality_weight"]

    # Portfolio global TCI
    portfolio_tci = (
        kpi["weighted_tci_portfolio_component"].sum() /
        (kpi["criticality_weight"].sum() if kpi["criticality_weight"].sum() else 1)
    )
    portfolio_summary = pd.DataFrame({
        "generated_at": [datetime.utcnow().isoformat()],
        "portfolio_tci": [portfolio_tci],
        "portfolio_tci_pct": [portfolio_tci / 5 * 100]
    })

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        products.to_excel(writer, sheet_name="Products", index=False)
        criteria.to_excel(writer, sheet_name="Criteria", index=False)
        scores.to_excel(writer, sheet_name="Scores", index=False)
        debt.to_excel(writer, sheet_name="DebtRegister", index=False)
        kpi.to_excel(writer, sheet_name="KPI", index=False)
        portfolio_summary.to_excel(writer, sheet_name="PortfolioSummary", index=False)

    wb = load_workbook(OUTPUT_FILE)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        format_sheet(ws)
    wb.save(OUTPUT_FILE)
    print(f"Excel generated: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()